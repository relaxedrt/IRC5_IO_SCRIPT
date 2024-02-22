import openpyxl

#Abrimos nuestro superexcel
ruta_superexcel = "SUPEREXCEL.xlsx"
superexcel = openpyxl.load_workbook(ruta_superexcel)

def main():
    #Nos dirijimos a la hoja del db en concreto
    n = input("Como se llama la hoja del db que quieres convertir?: ")
    hoja = superexcel[n]
    name = hoja.cell(row=10, column=3).value

    #Preguntamos si queremos que el db tenga el acceso optimizado
    print("")
    opt = input("Quieres que el db tenga el acceso optimizado? (y/n)")

    with open(f"{name}.db", "a") as db:
        #Comenzamos a escribir la cabecera del db
        db.write(f'DATA_BLOCK "{name}"\n')

        #Consultamos si tengrá el acceso optimizado
        if ((opt == "n")):
            db.write("{ S7_Optimized_Access := 'FALSE' }\n")
        else:
            db.write("{ S7_Optimized_Access := 'TRUE' }\n")

        db.write("VERSION : 0.1\n")
        db.write("NON_RETAIN\n")
        db.write("   VAR\n")

        #Entramos en un bucle para escribir las variables
        for i in range(13, 1048576):
            if (hoja.cell(row=i, column=3).value != None):
                db.write(f"      {hoja.cell(row=i, column=3).value} : {hoja.cell(row=i, column=4).value};\n")
            else:
                break
        
        #Escribimos el cierre de dbç
        db.write("   END_VAR\n")
        db.write("\n")
        db.write("\n")
        db.write("BEGIN\n")
        db.write("\n")
        db.write("END_DATA_BLOCK\n")

