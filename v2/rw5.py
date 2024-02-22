import openpyxl

#Accedemos al superexcel
name = "PLC_Project_Template_20240125_2.xlsx"
excel = openpyxl.load_workbook(name)

def main():
    #Preguntamos el nombre de la hoja a convertir
    n = input("Como se llama la hoja del db que quieres convertir?: ")
    hoja = excel[n]

    #Abrimos un archivo con el nombre del robot
    rob = hoja.cell(row=6, column=3).value
    file = open(str(rob) + ".txt", "w")
    for i in range(13, 1048576):
        #Determinamos el nombre de la variable
        varname = hoja.cell(row = i, column = 3).value

        #Determinamos el nombre de la unidad
        unittype = "PROFINET_SLAVE"

        #Determinamos el devicemap 
        #!!!!DE MOMENTO NO HAY ESTA CELDA POR LO QUE NOS LA INVENTAMOS!!!!!!!!
        #devicemap = hoja.cell(row = i, column = 8).value
        devicemap = "0"

        #Determinamos el comentario de la variable
        comm = hoja.cell(row = i, column = 6).value

        #Determinamos si es una entrada o una salida
        if hoja.cell(row = i, column = 7).value == "ROBIN":
            #Si es una entrada
            if hoja.cell(row = i, column = 4).value == "BOOL":
                vartype = "DI"
            elif  hoja.cell(row = i, column = 4).value == "BYTE":
                vartype = "GI"
        elif hoja.cell(row = i, column = 7).value == "ROBOUT":
            #Si es una entrada
            if hoja.cell(row = i, column = 4).value == "BOOL":
                vartype = "DO"
            elif  hoja.cell(row = i, column = 4).value == "BYTE":
                vartype = "GO"
    
    #Escribimos en el txt todo lo necesario para crear las variables
    file.write('      -Name "' + str(varname) + '" -SignalType "' + str(vartype) +'" -Unit "' + str(unittype) + '" -UnitMap "' + str(devicemap) + '"' + "\n")
