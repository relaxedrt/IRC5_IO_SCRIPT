from cProfile import label
import os
import openpyxl
nombre_archivo = input("Como quieres que se llame el archivo? ")
nombre_excel = input("Como se llama el excel? Recuerda que tiene que estar en la carpeta: ")
entradas_deseadas = input("Cuantas entradas vamos a convertir? ")
salidas_deseadas = input("Cuantas salidas vamos a convertir? ")
excel_document = openpyxl.load_workbook(nombre_excel + ".xlsx")
hoja = excel_document.active
file = open(str(nombre_archivo) + ".txt", "w")
for i in range(1, int(entradas_deseadas) + 1):
    nombre_variable = hoja.cell(row = 4 + i, column = 3)
    tipo_variable= hoja.cell(row = 4 + i, column = 8)
    unidad_variable = hoja.cell(row = 4 + i , column = 9)
    devicemap = hoja.cell(row = 4 + i , column = 7)
    coment = hoja.cell(row = 4 + i , column = 5)
    file.write('      -Name "' + str(nombre_variable.value) + '" -SignalType "' + str(tipo_variable.value) +'" -Unit "' + str(unidad_variable.value) + '" -UnitMap "' + str(devicemap.value) + '"' + os.linesep)
for n in range(1, int(salidas_deseadas) + 1):
    nombre_variable = hoja.cell(row = 4 + n, column = 12)
    tipo_variable= hoja.cell(row = 4 + n, column = 17)
    unidad_variable = hoja.cell(row = 4 + n , column = 18)
    devicemap = hoja.cell(row = 4 + n , column = 16)
    coment = hoja.cell(row = 4 + n , column = 14)
    file.write('      -Name "' + str(nombre_variable.value) + '" -SignalType "' + str(tipo_variable.value) +'" -Unit "' + str(unidad_variable.value) + '" -UnitMap "' + str(devicemap.value) + '"' + os.linesep)